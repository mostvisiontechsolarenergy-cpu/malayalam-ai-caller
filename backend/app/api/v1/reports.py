import io
import os
import uuid
from datetime import UTC, date, datetime, timedelta

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import func, select
from sqlalchemy.orm import Session


# Register Malayalam font
MALAYALAM_FONT = "FreeSans"
_font_registered = False


def _register_malayalam_font():
    global _font_registered
    if _font_registered:
        return
    font_paths = [
        "/usr/share/fonts/truetype/freefont/FreeSans.ttf",
        "/usr/share/fonts/truetype/freefont/FreeSansBold.ttf",
        "/usr/share/fonts/truetype/lohit-malayalam/Lohit-Malayalam.ttf",
        "/usr/share/fonts/truetype/lohit-devanagari/Lohit-Devanagari.ttf",
    ]
    for font_path in font_paths:
        if os.path.exists(font_path):
            try:
                pdfmetrics.registerFont(TTFont(MALAYALAM_FONT, font_path))
                _font_registered = True
                return
            except Exception:
                continue
    # Fallback: try to find any TTF font
    for root, dirs, files in os.walk("/usr/share/fonts"):
        for f in files:
            if f.endswith(".ttf"):
                try:
                    pdfmetrics.registerFont(TTFont(MALAYALAM_FONT, os.path.join(root, f)))
                    _font_registered = True
                    return
                except Exception:
                    continue

from app.core.dependencies import get_tenant_id, require_roles
from app.db.models import (
    AIAgent,
    AIConversation,
    AIConversationMessage,
    CallbackRequest,
    Client,
    PhoneCall,
    PhoneCallStatus,
    User,
    UserRole,
)
from app.db.session import get_db

router = APIRouter(prefix="/reports", tags=["reports"])


def _format_duration(seconds: int | None) -> str:
    if seconds is None:
        return "0s"
    minutes, secs = divmod(seconds, 60)
    if minutes > 0:
        return f"{minutes}m {secs}s"
    return f"{secs}s"


def _get_call_data(
    db: Session,
    company_id: uuid.UUID,
    start_date: date,
    end_date: date,
) -> list[dict]:
    """Fetch all calls for the company within the date range."""
    start_dt = datetime.combine(start_date, datetime.min.time()).replace(tzinfo=UTC)
    end_dt = datetime.combine(end_date + timedelta(days=1), datetime.min.time()).replace(
        tzinfo=UTC
    )

    calls = list(
        db.scalars(
            select(PhoneCall)
            .where(
                PhoneCall.company_id == company_id,
                PhoneCall.created_at >= start_dt,
                PhoneCall.created_at < end_dt,
            )
            .order_by(PhoneCall.created_at.desc())
        ).all()
    )

    results = []
    for call in calls:
        client = db.scalar(select(Client).where(Client.id == call.client_id))
        agent = db.scalar(select(AIAgent).where(AIAgent.id == call.agent_id))
        conversation = db.scalar(
            select(AIConversation).where(AIConversation.id == call.conversation_id)
        )

        # Get summary from conversation report
        summary = ""
        customer_request = ""
        if conversation and conversation.report_json:
            malayalam_report = conversation.report_json.get("malayalam_report", {})
            analysis = malayalam_report.get("analysis", {})
            if analysis:
                summary = analysis.get("summary_ml", "")
                customer_request = analysis.get("customer_requirement_ml", "")

        results.append(
            {
                "id": str(call.id),
                "client_name": client.name if client else "Unknown",
                "client_phone": call.destination,
                "agent_name": agent.name if agent else "Unknown",
                "status": call.status.value,
                "attended": call.status == PhoneCallStatus.COMPLETED,
                "duration_seconds": call.duration_seconds or 0,
                "duration_formatted": _format_duration(call.duration_seconds),
                "started_at": call.created_at.isoformat() if call.created_at else "",
                "summary": summary,
                "customer_request": customer_request,
            }
        )

    return results


def _generate_pdf(
    data: list[dict],
    start_date: date,
    end_date: date,
    company_name: str,
) -> io.BytesIO:
    """Generate PDF report."""
    _register_malayalam_font()
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=15 * mm,
        leftMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )

    styles = getSampleStyleSheet()
    
    # Custom styles for wrapping text with Malayalam font
    cell_style = ParagraphStyle(
        'CellStyle',
        parent=styles['Normal'],
        fontName=MALAYALAM_FONT,
        fontSize=8,
        leading=10,
        wordWrap='CJK',
    )
    header_cell_style = ParagraphStyle(
        'HeaderCellStyle',
        parent=styles['Normal'],
        fontName=MALAYALAM_FONT,
        fontSize=9,
        leading=11,
        textColor=colors.white,
        wordWrap='CJK',
    )
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Title'],
        fontName=MALAYALAM_FONT,
        fontSize=16,
        leading=20,
    )
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontName=MALAYALAM_FONT,
        fontSize=10,
        leading=14,
        spaceAfter=6,
    )
    stat_style = ParagraphStyle(
        'StatStyle',
        parent=styles['Normal'],
        fontName=MALAYALAM_FONT,
        fontSize=10,
        leading=14,
    )
    
    elements = []

    # Title
    elements.append(Paragraph(f"Daily Call Summary Report - {company_name}", title_style))
    elements.append(Paragraph(
        f"Period: {start_date.strftime('%d %b %Y')} to {end_date.strftime('%d %b %Y')}",
        subtitle_style
    ))
    elements.append(Spacer(1, 8 * mm))

    # Summary stats
    total_calls = len(data)
    completed_calls = sum(1 for d in data if d["attended"])
    total_duration = sum(d["duration_seconds"] for d in data)
    elements.append(Paragraph(f"<b>Total Calls:</b> {total_calls}", stat_style))
    elements.append(Paragraph(f"<b>Completed:</b> {completed_calls}", stat_style))
    elements.append(Paragraph(f"<b>Total Duration:</b> {_format_duration(total_duration)}", stat_style))
    elements.append(Spacer(1, 8 * mm))

    # Table with Paragraph cells for proper wrapping
    table_data = [[
        Paragraph("Phone Number", header_cell_style),
        Paragraph("Agent", header_cell_style),
        Paragraph("Attended", header_cell_style),
        Paragraph("Duration", header_cell_style),
        Paragraph("Summary", header_cell_style),
        Paragraph("Customer Request", header_cell_style),
    ]]
    for row in data:
        table_data.append([
            Paragraph(row["client_phone"], cell_style),
            Paragraph(row["agent_name"], cell_style),
            Paragraph("Yes" if row["attended"] else "No", cell_style),
            Paragraph(row["duration_formatted"], cell_style),
            Paragraph(row["summary"] or "—", cell_style),
            Paragraph(row["customer_request"] or "—", cell_style),
        ])

    col_widths = [70, 80, 50, 50, 200, 200]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a1a2e")),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f5f5")]),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    return buffer


def _generate_excel(
    data: list[dict],
    start_date: date,
    end_date: date,
) -> io.BytesIO:
    """Generate Excel report."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Call Report"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Title row
    ws.merge_cells("A1:F1")
    ws["A1"] = f"Daily Call Summary Report ({start_date.strftime('%d %b %Y')} - {end_date.strftime('%d %b %Y')})"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Summary row
    ws.merge_cells("A2:F2")
    total_calls = len(data)
    completed = sum(1 for d in data if d["attended"])
    total_dur = sum(d["duration_seconds"] for d in data)
    ws["A2"] = f"Total Calls: {total_calls} | Completed: {completed} | Total Duration: {_format_duration(total_dur)}"
    ws["A2"].font = Font(size=10)
    ws["A2"].alignment = Alignment(horizontal="center")

    # Headers
    headers = ["Phone Number", "Agent", "Attended", "Duration", "Summary", "Customer Request"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data rows
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    center_alignment = Alignment(horizontal="center", vertical="top")
    
    for row_idx, row_data in enumerate(data, 5):
        cell_a = ws.cell(row=row_idx, column=1, value=row_data["client_phone"])
        cell_a.border = thin_border
        cell_a.alignment = wrap_alignment
        
        cell_b = ws.cell(row=row_idx, column=2, value=row_data["agent_name"])
        cell_b.border = thin_border
        cell_b.alignment = wrap_alignment
        
        attended_cell = ws.cell(
            row=row_idx, column=3, value="Yes" if row_data["attended"] else "No"
        )
        attended_cell.border = thin_border
        attended_cell.alignment = center_alignment
        
        cell_d = ws.cell(row=row_idx, column=4, value=row_data["duration_formatted"])
        cell_d.border = thin_border
        cell_d.alignment = center_alignment
        
        cell_e = ws.cell(row=row_idx, column=5, value=row_data["summary"] or "—")
        cell_e.border = thin_border
        cell_e.alignment = wrap_alignment
        
        cell_f = ws.cell(row=row_idx, column=6, value=row_data["customer_request"] or "—")
        cell_f.border = thin_border
        cell_f.alignment = wrap_alignment

    # Column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 45
    ws.column_dimensions["F"].width = 45
    
    # Set row height for data rows to allow wrapping
    for row_idx in range(5, 5 + len(data)):
        ws.row_dimensions[row_idx].height = 45

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


@router.get("/call-summary")
def get_call_summary(
    start_date: date = Query(..., description="Start date (YYYY-MM-DD)"),
    end_date: date = Query(..., description="End date (YYYY-MM-DD)"),
    company_id: uuid.UUID = Depends(get_tenant_id),
    db: Session = Depends(get_db),
) -> dict:
    """Get call summary data for date range."""
    if start_date > end_date:
        raise HTTPException(status_code=400, detail="start_date must be before end_date")

    data = _get_call_data(db, company_id, start_date, end_date)

    total_calls = len(data)
    completed_calls = sum(1 for d in data if d["attended"])
    total_duration = sum(d["duration_seconds"] for d in data)

    return {
        "period": {
            "start": start_date.isoformat(),
            "end": end_date.isoformat(),
        },
        "summary": {
            "total_calls": total_calls,
            "completed_calls": completed_calls,
            "missed_calls": total_calls - completed_calls,
            "total_duration_seconds": total_duration,
            "total_duration_formatted": _format_duration(total_duration),
            "avg_duration_seconds": total_duration // total_calls if total_calls > 0 else 0,
        },
        "calls": data,
    }


@router.get("/call-summary/pdf")
def download_call_summary_pdf(
    start_date: date = Query(..., description="Start date (YYYY-MM-DD)"),
    end_date: date = Query(..., description="End date (YYYY-MM-DD)"),
    company_id: uuid.UUID = Depends(get_tenant_id),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    """Download call summary as PDF."""
    if start_date > end_date:
        raise HTTPException(status_code=400, detail="start_date must be before end_date")

    data = _get_call_data(db, company_id, start_date, end_date)

    # Get company name
    from app.db.models import Company

    company = db.scalar(select(Company).where(Company.id == company_id))
    company_name = company.name if company else "Company"

    pdf_buffer = _generate_pdf(data, start_date, end_date, company_name)

    filename = f"call-report-{start_date.isoformat()}-to-{end_date.isoformat()}.pdf"
    return StreamingResponse(
        pdf_buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/call-summary/excel")
def download_call_summary_excel(
    start_date: date = Query(..., description="Start date (YYYY-MM-DD)"),
    end_date: date = Query(..., description="End date (YYYY-MM-DD)"),
    company_id: uuid.UUID = Depends(get_tenant_id),
    db: Session = Depends(get_db),
) -> StreamingResponse:
    """Download call summary as Excel."""
    if start_date > end_date:
        raise HTTPException(status_code=400, detail="start_date must be before end_date")

    data = _get_call_data(db, company_id, start_date, end_date)

    excel_buffer = _generate_excel(data, start_date, end_date)

    filename = f"call-report-{start_date.isoformat()}-to-{end_date.isoformat()}.xlsx"
    return StreamingResponse(
        excel_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
