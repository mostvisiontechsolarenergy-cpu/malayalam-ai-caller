import csv
import io
import uuid
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path

from docx import Document as WordDocument
from openai import OpenAI
from openpyxl import load_workbook
from pypdf import PdfReader
from sqlalchemy import delete, select

from app.core.config import get_settings
from app.db.models import Document, DocumentChunk, DocumentStatus, EmbeddingStatus
from app.db.session import SessionLocal

ALLOWED_EXTENSIONS = {".pdf", ".docx", ".txt", ".csv", ".xlsx"}


@dataclass(frozen=True)
class ExtractedSection:
    text: str
    page_number: int | None = None
    section: str | None = None


def document_path(stored_name: str) -> Path:
    root = Path(get_settings().upload_dir).resolve()
    target = (root / stored_name).resolve()
    if root not in target.parents:
        raise ValueError("Unsafe document storage path")
    return target


def extract_sections(path: Path, file_type: str) -> list[ExtractedSection]:
    if file_type == ".pdf":
        reader = PdfReader(str(path))
        return [
            ExtractedSection(page.extract_text() or "", page_number=index + 1)
            for index, page in enumerate(reader.pages)
        ]
    if file_type == ".docx":
        document = WordDocument(str(path))
        sections: list[ExtractedSection] = []
        heading = None
        buffer: list[str] = []
        for paragraph in document.paragraphs:
            text = paragraph.text.strip()
            if not text:
                continue
            if paragraph.style and paragraph.style.name.startswith("Heading"):
                if buffer:
                    sections.append(ExtractedSection("\n".join(buffer), section=heading))
                heading, buffer = text, []
            else:
                buffer.append(text)
        if buffer:
            sections.append(ExtractedSection("\n".join(buffer), section=heading))
        return sections
    if file_type == ".txt":
        return [ExtractedSection(path.read_text(encoding="utf-8-sig"))]
    if file_type == ".csv":
        with path.open(encoding="utf-8-sig", newline="") as stream:
            rows = [" | ".join(cell.strip() for cell in row) for row in csv.reader(stream)]
        return [ExtractedSection("\n".join(rows), section="CSV data")]
    if file_type == ".xlsx":
        workbook = load_workbook(path, read_only=True, data_only=True)
        sections = []
        for sheet in workbook.worksheets:
            rows = [
                " | ".join("" if cell is None else str(cell) for cell in row)
                for row in sheet.iter_rows(values_only=True)
            ]
            sections.append(ExtractedSection("\n".join(rows), section=sheet.title))
        workbook.close()
        return sections
    raise ValueError("Unsupported document type")


def chunk_sections(
    sections: list[ExtractedSection], max_characters: int = 1200, overlap: int = 180
) -> list[tuple[str, ExtractedSection]]:
    chunks: list[tuple[str, ExtractedSection]] = []
    for section in sections:
        text = "\n".join(line.strip() for line in section.text.splitlines() if line.strip())
        start = 0
        while start < len(text):
            end = min(len(text), start + max_characters)
            if end < len(text):
                boundary = max(text.rfind("\n", start, end), text.rfind(". ", start, end))
                if boundary > start + max_characters // 2:
                    end = boundary + 1
            content = text[start:end].strip()
            if content:
                chunks.append((content, section))
            if end >= len(text):
                break
            start = max(start + 1, end - overlap)
    return chunks


def embed_texts(texts: list[str]) -> list[list[float]]:
    settings = get_settings()
    if not settings.openai_key_configured:
        return []
    client = OpenAI(api_key=settings.openai_api_key.get_secret_value())
    vectors: list[list[float]] = []
    for offset in range(0, len(texts), 50):
        response = client.embeddings.create(
            model=settings.openai_embedding_model,
            input=texts[offset : offset + 50],
            dimensions=settings.openai_embedding_dimensions,
            encoding_format="float",
        )
        vectors.extend(
            item.embedding for item in sorted(response.data, key=lambda item: item.index)
        )
    return vectors


def process_document(document_id: uuid.UUID) -> None:
    with SessionLocal() as db:
        document = db.scalar(select(Document).where(Document.id == document_id))
        if document is None:
            return
        document.status = DocumentStatus.PROCESSING
        document.embedding_status = EmbeddingStatus.PROCESSING
        document.error_message = None
        db.commit()
        try:
            sections = extract_sections(document_path(document.stored_name), document.file_type)
            chunks = chunk_sections(sections)
            if not chunks:
                raise ValueError("No readable text was found in this document")
            db.execute(delete(DocumentChunk).where(DocumentChunk.document_id == document.id))
            records = [
                DocumentChunk(
                    company_id=document.company_id,
                    document_id=document.id,
                    content=content,
                    metadata_json={"filename": document.filename},
                    page_number=section.page_number,
                    section=section.section,
                    token_estimate=max(1, len(content) // 4),
                    chunk_index=index,
                )
                for index, (content, section) in enumerate(chunks)
            ]
            db.add_all(records)
            document.status = DocumentStatus.READY
            document.extracted_characters = sum(len(record.content) for record in records)
            document.chunk_count = len(records)
            document.processed_at = datetime.now(UTC)
            settings = get_settings()
            if not settings.openai_key_configured:
                document.embedding_status = EmbeddingStatus.SKIPPED_NO_KEY
                document.embedding_model = None
                db.commit()
                return
            document.embedding_status = EmbeddingStatus.PROCESSING
            document.embedding_model = settings.openai_embedding_model
            db.commit()
        except Exception as exc:  # extraction failure must be visible to the administrator
            db.rollback()
            document = db.scalar(select(Document).where(Document.id == document_id))
            if document:
                document.status = DocumentStatus.FAILED
                document.embedding_status = EmbeddingStatus.FAILED
                document.error_message = str(exc)[:1000]
                document.processed_at = datetime.now(UTC)
                db.commit()
            return

        try:
            vectors = embed_texts([record.content for record in records])
            if len(vectors) != len(records):
                raise RuntimeError("Embedding provider returned an incomplete result")
            for record, vector in zip(records, vectors, strict=True):
                record.embedding = vector
            document.embedding_status = EmbeddingStatus.READY
            document.error_message = None
            db.commit()
        except Exception as exc:  # extracted text remains usable through lexical retrieval
            db.rollback()
            document = db.scalar(select(Document).where(Document.id == document_id))
            if document:
                document.status = DocumentStatus.READY
                document.embedding_status = EmbeddingStatus.FAILED
                document.error_message = f"Embedding failed: {exc}"[:1000]
                db.commit()


def extract_text_for_testing(data: bytes, file_type: str) -> str:
    """Small pure helper used by ingestion tests without persistent storage."""
    if file_type == ".txt":
        return data.decode("utf-8-sig")
    if file_type == ".csv":
        rows = csv.reader(io.StringIO(data.decode("utf-8-sig")))
        return "\n".join(" | ".join(row) for row in rows)
    raise ValueError("Use a file-backed extractor for this format")
